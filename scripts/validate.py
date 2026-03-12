"""Validate canonical data, CSV exports, and workbook outputs."""

from __future__ import annotations

import sys

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - dependency guard
    raise SystemExit(f"Missing dependency: {exc}. Run: uv sync")

try:
    from export_csv import (
        TRIALS_LABELED_COLUMNS,
        build_curation_audit_rows,
        build_missing_requirements_rows,
        build_trial_rules_rows,
        build_trials_labeled_rows,
        summarize_rules_by_trial,
    )
    from generate_metrics import build_metrics
    from load_data import (
        load_condition_hits,
        load_criterion_catalog,
        load_csv_output,
        load_dataset_metadata,
        load_fields,
        load_memberships,
        load_metrics,
        load_not_evaluable,
        load_raw_trials,
        load_review_overrides,
        load_trial_rules,
        load_trials,
    )
    from pipeline_utils import KNOWN_NOISY_NCTS, MAPPED_CATEGORY_COUNT, SEED_CONDITION_COUNT, OUTPUTS
except ModuleNotFoundError:  # pragma: no cover - package import path
    from scripts.export_csv import (
        TRIALS_LABELED_COLUMNS,
        build_curation_audit_rows,
        build_missing_requirements_rows,
        build_trial_rules_rows,
        build_trials_labeled_rows,
        summarize_rules_by_trial,
    )
    from scripts.generate_metrics import build_metrics
    from scripts.load_data import (
        load_condition_hits,
        load_criterion_catalog,
        load_csv_output,
        load_dataset_metadata,
        load_fields,
        load_memberships,
        load_metrics,
        load_not_evaluable,
        load_raw_trials,
        load_review_overrides,
        load_trial_rules,
        load_trials,
    )
    from scripts.pipeline_utils import KNOWN_NOISY_NCTS, MAPPED_CATEGORY_COUNT, SEED_CONDITION_COUNT, OUTPUTS

PASS = "OK"
FAIL = "FAIL"


def check(condition: bool, message: str, errors: list[str]) -> None:
    if condition:
        print(f"  {PASS} {message}")
    else:
        print(f"  {FAIL} {message}")
        errors.append(message)


def stringify_row(row: dict, columns: list[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    for column in columns:
        value = row.get(column, "")
        if value is None:
            result[column] = ""
        else:
            result[column] = str(value)
    return result


def workbook_summary_map(path) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Summary"]
    values: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        value = row[1] if len(row) > 1 else ""
        if key:
            values[str(key)] = "" if value is None else str(value)
    return values


def run_validation() -> dict:
    errors: list[str] = []
    print("=" * 60)
    print("Validating pipeline outputs")
    print("=" * 60)

    print("\n[1] Loading artifacts...")
    trials = load_trials()
    raw_trials = load_raw_trials()
    raw_hits = load_condition_hits()
    memberships = load_memberships()
    fields = load_fields()
    dataset = load_dataset_metadata()
    catalog = load_criterion_catalog()
    rules = load_trial_rules()
    overrides = load_review_overrides()
    not_evaluable = load_not_evaluable()
    metrics = load_metrics()
    trials_labeled = load_csv_output("trials_labeled.csv")
    trial_rules_csv = load_csv_output("trial_rules.csv")
    missing_by_trial_csv = load_csv_output("missing_requirements_by_trial.csv")
    missing_summary_csv = load_csv_output("missing_requirements_summary.csv")
    curation_audit_csv = load_csv_output("curation_audit.csv")
    check(True, "All canonical JSON and CSV outputs loaded", errors)

    print("\n[2] iDHEA sync outputs...")
    for required_key in [
        "dataset_title",
        "dataset_slug",
        "source_url",
        "page_last_updated",
        "synced_at",
        "metrics",
        "field_count",
    ]:
        check(bool(dataset.get(required_key)), f"Dataset metadata includes {required_key}", errors)
    for row in fields[:5]:
        for provenance_key in ["source_url", "source_section", "page_last_updated", "synced_at"]:
            check(bool(row.get(provenance_key)), f"Field {row['field_name']} includes {provenance_key}", errors)

    print("\n[3] Referential integrity...")
    trial_ids = {trial["nct_id"] for trial in trials}
    catalog_ids = {row["criterion_id"] for row in catalog}
    membership_orphans = [row for row in memberships if row["nct_id"] not in trial_ids]
    check(not membership_orphans, "All memberships reference curated trials", errors)
    rule_orphans = [row for row in rules if row["nct_id"] not in trial_ids]
    check(not rule_orphans, "All trial rules reference curated trials", errors)
    bad_criterion_refs = [row for row in rules if row["criterion_id"] not in catalog_ids]
    check(not bad_criterion_refs, "All trial rules reference criterion catalog IDs", errors)
    leaked_noisy = [row for row in memberships if row["nct_id"] in KNOWN_NOISY_NCTS]
    check(not leaked_noisy, "Known noisy trials are excluded from curated memberships", errors)

    print("\n[4] Metrics regeneration...")
    regenerated_metrics = build_metrics(trials, memberships, rules)
    for key in [
        "unique_trials_total",
        "condition_memberships_total",
        "recruiting_now_total",
        "pipeline_open_total",
        "active_total",
        "mapped_trials_total",
        "verified_mapped_trials_total",
    ]:
        check(
            regenerated_metrics[key] == metrics.get(key),
            f"metrics.json {key} matches regenerated value",
            errors,
        )

    print("\n[5] CSV determinism...")
    known_missing_fields = {row["field_name"] for row in not_evaluable}
    rule_summaries = summarize_rules_by_trial(rules, known_missing_fields)
    expected_trials_labeled = [
        stringify_row(row, TRIALS_LABELED_COLUMNS)
        for row in build_trials_labeled_rows(trials, memberships, rule_summaries)
    ]
    check(expected_trials_labeled == trials_labeled, "trials_labeled.csv is deterministic", errors)

    trial_lookup = {trial["nct_id"]: trial for trial in trials}
    expected_trial_rules = [
        stringify_row(
            row,
            [
                "mapping_id",
                "nct_id",
                "title",
                "status",
                "condition_category",
                "criterion_id",
                "criterion_type",
                "criterion_text_original",
                "operator",
                "value",
                "unit",
                "idhea_fields",
                "external_dependencies",
                "confidence",
                "manual_review_required",
                "human_verified",
                "extraction_method",
                "model_name",
                "evidence_excerpt",
                "reasoning",
                "evidence_url",
            ],
        )
        for row in build_trial_rules_rows(rules, trial_lookup)
    ]
    check(expected_trial_rules == trial_rules_csv, "trial_rules.csv is deterministic", errors)

    expected_missing_by_trial, expected_missing_summary = build_missing_requirements_rows(
        rules, trial_lookup, memberships
    )
    expected_missing_by_trial_rows = [
        stringify_row(
            row,
            [
                "nct_id",
                "title",
                "status",
                "primary_condition",
                "condition_category",
                "missing_dependency",
                "source_rule_ids",
                "source_url",
            ],
        )
        for row in expected_missing_by_trial
    ]
    expected_missing_summary_rows = [
        stringify_row(row, ["missing_dependency", "condition_category", "status", "trial_count"])
        for row in expected_missing_summary
    ]
    check(
        expected_missing_by_trial_rows == missing_by_trial_csv,
        "missing_requirements_by_trial.csv is deterministic",
        errors,
    )
    check(
        expected_missing_summary_rows == missing_summary_csv,
        "missing_requirements_summary.csv is deterministic",
        errors,
    )
    expected_audit_rows = [
        stringify_row(
            row,
            [
                "nct_id",
                "condition_category",
                "condition_label",
                "condition_query",
                "title",
                "status",
                "source_url",
                "decision",
                "reason",
                "override_source",
                "corrected_conditions",
            ],
        )
        for row in build_curation_audit_rows(raw_hits, memberships, overrides)
    ]
    check(expected_audit_rows == curation_audit_csv, "curation_audit.csv is deterministic", errors)

    print("\n[6] Workbook summary...")
    workbook_path = OUTPUTS / "trial_prescreening_qa.xlsx"
    check(workbook_path.exists(), "Workbook exists", errors)
    if workbook_path.exists():
        summary_map = workbook_summary_map(workbook_path)
        expected_summary_pairs = {
            "Unique trials total": str(metrics["unique_trials_total"]),
            "Condition memberships total": str(metrics["condition_memberships_total"]),
            "Recruiting now (RECRUITING only)": str(metrics["recruiting_now_total"]),
            "Pipeline open (RECRUITING + NOT_YET_RECRUITING + ENROLLING_BY_INVITATION)": str(
                metrics["pipeline_open_total"]
            ),
            "Active total": str(metrics["active_total"]),
            "Mapped trials total": str(metrics["mapped_trials_total"]),
            "Verified mapped trials total": str(metrics["verified_mapped_trials_total"]),
        }
        for label, value in expected_summary_pairs.items():
            check(summary_map.get(label) == value, f"Workbook summary matches {label}", errors)

    print("\n[7] Cross-artifact consistency...")
    # Condition category count must match MAPPED_CATEGORY_COUNT
    unique_conditions = {row["condition_category"] for row in memberships}
    check(
        len(unique_conditions) == MAPPED_CATEGORY_COUNT,
        f"Condition categories ({len(unique_conditions)}) matches MAPPED_CATEGORY_COUNT ({MAPPED_CATEGORY_COUNT})",
        errors,
    )

    # Gap counts: no summary/total row should be counted as a trial
    # Each gap dependency should count unique trials only
    from collections import defaultdict

    dep_trials: dict[str, set[str]] = defaultdict(set)
    for rule in rules:
        if rule["confidence"] == "not_evaluable":
            for dep in rule.get("external_dependencies", []):
                dep_trials[dep].add(rule["nct_id"])
    for dep, trial_set in dep_trials.items():
        check(
            all(nct_id.startswith("NCT") for nct_id in trial_set),
            f"Gap '{dep}' has only valid NCT IDs (no summary rows counted)",
            errors,
        )

    print("\n" + "=" * 60)
    if errors:
        print(f"VALIDATION FAILED: {len(errors)} issue(s)")
        for issue in errors:
            print(f"  - {issue}")
        raise SystemExit(1)

    print("VALIDATION PASSED")
    return {
        "trials": len(trials),
        "memberships": len(memberships),
        "rules": len(rules),
        "fields": len(fields),
        "raw_trials": len(raw_trials),
    }


if __name__ == "__main__":
    try:
        result = run_validation()
        print(result)
    except Exception as exc:  # pragma: no cover - CLI path
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
