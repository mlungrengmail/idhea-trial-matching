"""Generate the QA workbook from canonical metrics and CSV exports.

Output:
  outputs/trial_prescreening_qa.xlsx
"""

from __future__ import annotations

import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:  # pragma: no cover - dependency guard
    raise SystemExit(f"Missing dependency: {exc}. Run: uv sync")

try:
    from load_data import (
        load_csv_output,
        load_dataset_metadata,
        load_fields,
        load_metrics,
        load_not_evaluable,
    )
    from pipeline_utils import OUTPUTS
except ModuleNotFoundError:  # pragma: no cover - package import path
    from scripts.load_data import (
        load_csv_output,
        load_dataset_metadata,
        load_fields,
        load_metrics,
        load_not_evaluable,
    )
    from scripts.pipeline_utils import OUTPUTS

HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
ALT_FILL = PatternFill(start_color="F6F9FC", end_color="F6F9FC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header(ws, row: int, width: int) -> None:
    for col in range(1, width + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_body(ws, start_row: int, end_row: int, width: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, width + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if (row - start_row) % 2 == 1:
                cell.fill = ALT_FILL


def autofit_columns(ws) -> None:
    for column_cells in ws.columns:
        values = [len(str(cell.value or "")) for cell in column_cells[:200]]
        width = max(values, default=12)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 60)


def append_table(ws, headers: list[str], rows: list[dict]) -> None:
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    if rows:
        style_body(ws, 2, len(rows) + 1, len(headers))
    autofit_columns(ws)


def generate() -> None:
    metrics = load_metrics()
    dataset = load_dataset_metadata()
    fields = load_fields()
    not_evaluable = load_not_evaluable()
    trials_labeled = load_csv_output("trials_labeled.csv")
    trial_rules = load_csv_output("trial_rules.csv")
    missing_by_trial = load_csv_output("missing_requirements_by_trial.csv")
    missing_summary = load_csv_output("missing_requirements_summary.csv")
    curation_audit = load_csv_output("curation_audit.csv")

    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    style_header(summary, 1, 2)
    summary_rows = [
        {"Metric": "Generated At", "Value": metrics["generated_at"]},
        {"Metric": "Dataset", "Value": dataset["dataset_title"]},
        {"Metric": "Dataset Page Last Updated", "Value": dataset["page_last_updated"]},
        {"Metric": "Unique trials total", "Value": metrics["unique_trials_total"]},
        {"Metric": "Condition memberships total", "Value": metrics["condition_memberships_total"]},
        {"Metric": "Recruiting now (RECRUITING only)", "Value": metrics["recruiting_now_total"]},
        {
            "Metric": "Pipeline open (RECRUITING + NOT_YET_RECRUITING + ENROLLING_BY_INVITATION)",
            "Value": metrics["pipeline_open_total"],
        },
        {"Metric": "Active total", "Value": metrics["active_total"]},
        {"Metric": "Mapped trials total", "Value": metrics["mapped_trials_total"]},
        {"Metric": "Verified mapped trials total", "Value": metrics["verified_mapped_trials_total"]},
    ]
    for row in summary_rows:
        summary.append([row["Metric"], row["Value"]])
    condition_counts = metrics.get("condition_counts", {})
    summary.append(["", ""])
    summary.append(["Condition", "Trial Count"])
    style_header(summary, len(summary_rows) + 3, 2)
    for key, value in condition_counts.items():
        summary.append([key, value])
    style_body(summary, 2, summary.max_row, 2)
    autofit_columns(summary)

    dataset_ws = wb.create_sheet("Dataset Metadata")
    dataset_rows = [
        {"Key": "dataset_title", "Value": dataset.get("dataset_title", "")},
        {"Key": "dataset_slug", "Value": dataset.get("dataset_slug", "")},
        {"Key": "source_url", "Value": dataset.get("source_url", "")},
        {"Key": "page_last_updated", "Value": dataset.get("page_last_updated", "")},
        {"Key": "api_last_updated", "Value": dataset.get("api_last_updated", "")},
        {"Key": "field_count", "Value": dataset.get("field_count", "")},
        {"Key": "update_info", "Value": dataset.get("update_info", "")},
        {"Key": "tags", "Value": "; ".join(dataset.get("tags", []))},
        {"Key": "content_summary", "Value": dataset.get("content_summary", "")},
    ]
    append_table(dataset_ws, ["Key", "Value"], dataset_rows)

    fields_ws = wb.create_sheet("iDHEA Fields")
    append_table(
        fields_ws,
        [
            "field_name",
            "source_label",
            "definition",
            "units",
            "modality",
            "source_section",
            "source_subsection",
            "page_last_updated",
        ],
        fields,
    )

    trials_ws = wb.create_sheet("Trials Labeled")
    append_table(trials_ws, list(trials_labeled[0].keys()) if trials_labeled else [], trials_labeled)

    rules_ws = wb.create_sheet("Trial Rules")
    append_table(rules_ws, list(trial_rules[0].keys()) if trial_rules else [], trial_rules)

    missing_ws = wb.create_sheet("Missing By Trial")
    append_table(
        missing_ws,
        list(missing_by_trial[0].keys()) if missing_by_trial else [],
        missing_by_trial,
    )

    missing_summary_ws = wb.create_sheet("Missing Summary")
    append_table(
        missing_summary_ws,
        list(missing_summary[0].keys()) if missing_summary else [],
        missing_summary,
    )

    not_eval_ws = wb.create_sheet("Missing Field Catalog")
    append_table(
        not_eval_ws,
        ["field_name", "description", "trial_prevalence", "impact", "remediation"],
        not_evaluable,
    )

    audit_ws = wb.create_sheet("Curation Audit")
    append_table(audit_ws, list(curation_audit[0].keys()) if curation_audit else [], curation_audit)

    out_path = OUTPUTS / "trial_prescreening_qa.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    try:
        generate()
    except Exception as exc:  # pragma: no cover - CLI path
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
